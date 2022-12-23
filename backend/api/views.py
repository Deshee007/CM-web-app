import io
from rest_framework.response import Response
from django.http import FileResponse, HttpResponse, HttpResponseBadRequest, HttpResponseRedirect, JsonResponse
from rest_framework.decorators import api_view
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment,PatternFill
from base.models import Course,Discipline,Learning_Units,Learning_Outcomes,Study_Materials,CurriculumFiles
from .serializers import CourseSerializer, DisciplineSerializer, Learning_UnitsSerializer,Learning_OutcomesSerializer,Study_MaterialsSerializer,CurriculumFilesSerializer
from openpyxl import load_workbook
import json

def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def testMerge(row, column,sheet):
    cell = sheet.cell(row, column)
    for mergedCell in sheet.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            return True
    return False


def FetchAllData(file):                 ## Fetch all the data from the excel sheet 
    try:
        workbook = load_workbook(io.BytesIO(file))            # Load the excel sheet
    except:
        workbook = load_workbook(file)
    worksheet = workbook.active                      # Select the active one 

    mc = worksheet.max_column                  # Finding maximum columns and storing the number in mc 
    starting_character= 'A'            # Starting column  -- B
    TotalColumns = 0
    for i in range(1,mc):
        char = ord(starting_character)          # Taking the order of the B
        location = starting_character + str(16)         # Adding 15 to B and storing in location
        # print(ws[location])
        if worksheet[location].value == None: 
            break 
        val = char + 1 
        starting_character = chr(val)
        TotalColumns +=1

    # print("Total Columns: ",TotalColumns)
    mc= TotalColumns
    programs = []
    list_of_programs = []
            
    merge_columns = 0
    num =1
    prev_key = ""
    for i in range(1,mc+1):
        program = worksheet.cell (row = 15, column = i)         # Taking the value of the row and column and storing in the programs 
        if program.value == None:
            merge_columns +=1
        else:
            if i == 1:
                pass
            elif i==2:
                prev_key = program.value
                
            else:   
                programs.append( {
                    'name': prev_key,
                    'merge_columns': merge_columns + 1 
                })
                prev_key = program.value
                merge_columns = 0
                num+=1
    programs.append( {
        'name': prev_key,
        'merge_columns': merge_columns + 1 
    })
    list_of_programs = {"total_programs": num, "data":programs}
    # print(list_of_programs)


    modules =[]
    for i in range(2,mc+1):
        module = worksheet.cell (row = 16, column = i)         # Taking the value of the row and column and storing in the programs 
        # print(module.value)
        mod = module.value.split('\n')
        modules.append( {
            'module_code': mod[0],
            'module_name': mod[1]
        })
       
        # print(module.value)
        
    # print (modules)




    # FOr fetching data inside table 

    # print("Total Columns: ",mc+1)
    mr = worksheet.max_row                     # Finding maximum rows and storing the number in mr 

    final_data = [
        {"table_data": []},
        {"list_programs": []},
        {"modules": []}
    ]                            # Create an object where we can store the data 

    final_key = ""                          # This is a final key which will store the merge columns 
    all_data = []                           # This is a dict in which each final key data will be stored 
    each_key = ""                           # Single key with each row column 1 will be stored 


    for i in range(17,mr-1):            # Loop through the rows starting from 17 
        data = []                       # In this array data for each row will be stored after each key 
        keyvalue = False                 # To check if the data is the final key or not 
        # final_data = [
        #     {"table_data": [
                    
        #                 {
        #                     "title": "knowledge",
        #                     "data": []
        #                 }, 
        #                 {
        #                     "title": "intelligent",
        #                     "data": []
        #                 }, 
                    
        #     ]}
        # ]
        for j in range ( 1 , mc + 1 ):      # Loop through columns starting from 1 
          
            # read cell value from Excel source file    
            c = worksheet.cell (row = i, column = j)         # Taking the value of the row and column and storing in the variable c 

            if testMerge(i,j,worksheet):                   # If its final key then it will make keyvalue to True so that it will save it in final key 
           
                if i!=17:                           # If row is not first one 
                    final_data[0]["table_data"].append({"title": final_key,"data": all_data})         # store all data in final data dict and with final key to be the final key 
                    
                keyvalue = True                                 # Common: Keyvalue = True to show that it is key 
                final_key = c.value
                # print(final_key)
                                                    # making inde = 1 to start from 1 again to store data for each row 
                all_data = []                                  # All data to store all rows of this type 
                break
            else:
                if j == 1:
                    each_key = c.value 
                else:
                    keyvalue = False
                    data.append(c.value)
                   
        if keyvalue == False:
            all_data.append( {
                "column_name": each_key,
                "data": data
            })
         

    final_data[0]["table_data"].append({"title": final_key,"data": all_data}) 

    final_data[1]['list_programs'] = list_of_programs
    final_data[2]['modules'] = modules
    return final_data


#  Get all data 

@api_view(['GET'])
def getAllFiles(request):
    cFiles = CurriculumFiles.objects.all()
    # for file in cFiles:
    #     print(file.excelFile)
    # print(cFiles)
    serializer  = CurriculumFilesSerializer(cFiles,many = True)
    return Response(serializer.data)

@api_view(['GET'])
def download(request):
    id= request.query_params['id']
 
    cFile = CurriculumFiles.objects.get(id=id)

    file = cFile.excelFile
    

    # get_binary = obj.image
    if file is None:
        return JsonResponse({'status_message': 'Resource does not contian image'})
    if isinstance(file, memoryview):
        binary_io = io.BytesIO(file.tobytes())
    else:
        binary_io = io.BytesIO(file)
    response = FileResponse(binary_io)
    response['Content-Type'] = 'application/x-binary'
    response['Content-Disposition'] = 'attachment; filename="{}"'.format(cFile.filename) # Can set custom filename.
    return response

@api_view(['GET'])
def remove(request):
    id= request.query_params['id']
    cFile = CurriculumFiles.objects.get(id=id)
    if cFile is None:
        return Response({
            "status": 400,
            "message": "No File Found"
        })
    cFile.delete()
    return Response({
        "status": 200,
        "message": "Successfully deleted"
    })

@api_view(['GET'])
def deleteFiles(request):
    cFiles = CurriculumFiles.objects.all()
    cFiles.delete()
    return Response(200)


@api_view(['GET'])
def getDisciplines(request):
    disciplines = Discipline.objects.all()
    serializer  = DisciplineSerializer(disciplines,many = True)
    # print(serializer.data)
    
    return Response(serializer.data)


@api_view(['GET'])
def getCourses(request):
    courses = Course.objects.all()
    serializer  = CourseSerializer(courses,many = True)

    final_data=  []
    for eachData in serializer.data:
        # print(eachData)
        dip_data = Discipline.objects.get(id=eachData['discipline_id'])
        eachData['discipline_code'] = dip_data.discipline_code
        final_data.append(eachData)

    return Response(final_data)

@api_view(['GET'])
def getLearningUnits(request):
    learning_units = Learning_Units.objects.all()

    serializer  = Learning_UnitsSerializer(learning_units,many = True)

    final_data=  []
    for eachData in serializer.data:
        # print(eachData)
        dip_data = Course.objects.get(id=eachData['course_id'])
        eachData['course_code'] = dip_data.course_code
        final_data.append(eachData)


    return Response(final_data)
    
@api_view(['GET'])
def getLearningOutcomes(request):
    learning_outcomes = Learning_Outcomes.objects.all()
    serializer  = Learning_OutcomesSerializer(learning_outcomes,many = True)
    return Response(serializer.data)
   
@api_view(['GET'])
def getStudyMaterials(request):
    study_materials = Study_Materials.objects.all()
    serializer  = Study_MaterialsSerializer(study_materials,many = True)
    return Response(serializer.data)


#  Get Data by ID 
@api_view(['GET'])
def getDisciplineByID(request):
    id = request.GET["id"]
    # print(id)
    # data["id"]=int(data["id"])
    disciplines = Discipline.objects.get(id=id)
    serializer  = DisciplineSerializer(disciplines,many = False)
    return Response(serializer.data)
    

@api_view(['GET'])
def getCourseByID(request):
    id = request.GET["id"]

    courses = Course.objects.get(id=id)
    serializer  = CourseSerializer(courses,many = False)
    return Response(serializer.data)

@api_view(['GET'])
def getLearningUnitByID(request):
    id = request.GET["id"]

    learning_units = Learning_Units.objects.get(id=id)
    serializer  = Learning_UnitsSerializer(learning_units,many = False)
    return Response(serializer.data)
    
@api_view(['GET'])
def getLearningOutcomeByID(request):
    id = request.GET["id"]
    learning_outcomes = Learning_Outcomes.objects.get(id=id)
    serializer  = Learning_OutcomesSerializer(learning_outcomes,many = False)
    return Response(serializer.data)
   
@api_view(['GET'])
def getStudyMaterialByID(request):
    id = request.GET["id"]
    study_materials = Study_Materials.objects.get(id=id)
    serializer  = Study_MaterialsSerializer(study_materials,many = False)
    return Response(serializer.data)


# Adding Data
@api_view(['POST'])
def addDiscipline(request):
    serializer = DisciplineSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
    else:
        return HttpResponseBadRequest(
            "Discipline with that code already exist" 
        )
    return Response(serializer.data)


@api_view(['POST'])
def addCourse(request):
    serializer = CourseSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
    else:
        return HttpResponseBadRequest(
            "Course with that code already exist" 
        )
    return Response(serializer.data)

@api_view(['POST'])
def addLearningUnits(request):
    serializer = Learning_UnitsSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
    
    return Response(serializer.data)

@api_view(['POST'])
def addLearningOutcomes(request):
    serializer = Learning_OutcomesSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
    
    return Response(serializer.data)

@api_view(['POST'])
def addStudyMaterials(request):
    
    serializer = Study_MaterialsSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()
    
    return Response(serializer.data)




# Saving file in database

@api_view(['POST'])
def save(request):
    try:

        # print("File is sent")
        # request.data['excelFile'] = request.data['excelFile']
     
        newData = request.data
        # print(newData)
        cFiles = CurriculumFiles.objects.create(filename = request.data['filename'], excelFile = request.data['excelFile'].read())
        cFiles.save()
        return HttpResponse(200)
    except:
        return HttpResponseBadRequest(400)
  




#  Updating Data
@api_view(['POST'])
def updateDiscipline(request):
    data = request.data
    # print(request.data)
    disciplines = Discipline.objects.get(id=data["id"])
    disciplines.discipline_name = data["discipline_name"]
    disciplines.save()
    return Response({
        "status": 200,
        "message": "Successfully updated discipline"
    })
@api_view(['POST'])
def updateCourse(request):
    data = request.data
    course = Course.objects.get(id=data["id"])
    did = data["discipline_id"]

    discipline_id = Discipline.objects.get(id = did)
    course.course_name = data["course_name"]
    course.discipline_id = discipline_id
    course.save()
    return Response({
        "status": 200,
        "message": "Successfully updated course"
    })

@api_view(['POST'])
def updateLearningUnits(request):
    data = request.data
    learning_units = Learning_Units.objects.get(id=data["id"])
    cid = data["course_id"]
    course_id = Course.objects.get(id = cid)
    learning_units.title = data["title"]
    learning_units.level = data["level"]
    learning_units.course_id = course_id
    learning_units.save()
    return Response({
        "status": 200,
        "message": "Successfully updated Learning Unit"
    })

@api_view(['POST'])
def updateLearningOutcomes(request):
    data = request.data
    learning_outcomes = Learning_Outcomes.objects.get(id=data["id"])
    learning_outcomes.title = data["title"]
    learning_outcomes.state = data["state"]
    learning_outcomes.save()
    return Response({
        "status": 200,
        "message": "Successfully updated Learning Outcome"
    })

@api_view(['POST'])
def updateStudyMaterials(request):
    data = request.data
    study_materials = Study_Materials.objects.get(id=data["id"])
    study_materials.title = data["title"]
    study_materials.material_type = data["material_type"]
    study_materials.save()
    return Response({
        "status": 200,
        "message": "Successfully updated Study Material"
    })


# Deleting Data
@api_view(['POST'])
def deleteDiscipline(request):
  data= request.data
  disciplines = Discipline.objects.get(id=data["id"])
  disciplines.delete()
  return Response({
    "status": 200,
    "message": "Successfully deleted"
  })

@api_view(['POST'])
def deleteCourse(request):
  data= request.data
  course = Course.objects.get(id=data["id"])
  course.delete()
  return Response({
    "status": 200,
    "message": "Successfully deleted"
  })
  
@api_view(['POST'])
def deleteLearningUnits(request):
  data= request.data
  learning_units = Learning_Units.objects.get(id=data["id"])
  learning_units.delete()
  return Response({
    "status": 200,
    "message": "Successfully deleted"
  })

@api_view(['POST'])
def deleteLearningOutcomes(request):
  data= request.data
  learning_outcomes = Learning_Outcomes.objects.get(id=data["id"])
  learning_outcomes.delete()
  return Response({
    "status": 200,
    "message": "Successfully deleted"
  })

@api_view(['POST'])
def deleteStudyMaterials(request):
  data= request.data
  study_materials = Study_Materials.objects.get(id=data["id"])
  study_materials.delete()
  return Response({
    "status": 200,
    "message": "Successfully deleted"
  })



@api_view (['POST'])
def getData(request):
    filename = request.FILES['file'].file
    # filename = request.FILES['file'].file
    # filename = os.path.abspath("../Curriculum-map-template.xlsx")
    data = FetchAllData(filename)
    return Response(data)

@api_view(['GET'])
def getFileAndData(request):
    id= request.query_params['id']
    # print(id)
    cFile = CurriculumFiles.objects.get(id=id)
    if cFile is None:
        pass
    else:

        filename = cFile.excelFile
        # print(filename)
        filename = filename
        # filename = os.path.abspath("../Curriculum-map-template.xlsx")
        data = FetchAllData(filename)
        return Response(
            data
            )
    return HttpResponseBadRequest()

@api_view(['POST'])
def createData(request):
    all_data = request.data
    all_data = str(request.data)
    all_data = all_data.replace("'",'"')
    all_data = json.loads(all_data)

    savingFileName = all_data["fileName"]
    print(all_data)

    TotalColumns = 0 

    # all_data = {
    #     "programmes": [{"merge_columns": 3, "name": "Semester 1"},{"merge_columns": 2, "name": "Semester 2"},{"merge_columns": 4, "name": "Semester 3"},{"merge_columns": 3, "name": "Semester 4"}],
    #     "modules": [{"module_name": "Module 1", "module_code": "1"},{"module_name": "Module 2", "module_code": "2"},{"module_name": "Module 3", "module_code": "3"},
    #                 {"module_name": "Module 4", "module_code": "4"},{"module_name": "Module 5", "module_code": "5"},{"module_name": "Module 6", "module_code": "6"},{"module_name": "Module 7", "module_code": "7"},
    #                 {"module_name": "Module 8", "module_code": "8"},{"module_name": "Module 9", "module_code": "9"},{"module_name": "Module 10", "module_code": "10"},{"module_name": "Module 11", "module_code": "11"},{"module_name": "Module 12", "module_code": "12"}],
    #     "table_data": [
    #         {"title": "heading 1", "data":[
    #                 {"heading": "1", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]},
    #                 {"heading": "2", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]},
    #                 {"heading": "3", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]},
    #             ]
    #         },
    #         {"title": "heading 2", "data":[
    #                 {"heading": "1", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]},
    #                 {"heading": "2", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]}
    #             ]
    #         },
    #         {"title": "heading 3", "data":[
    #                 {"heading": "1", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]},
    #                 {"heading": "2", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]},
    #                 {"heading": "3", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]},
    #                 {"heading": "4", "data":["1","2","3","4","5","6","7","8","9","10","11","12"]},

    #             ]
    #         },
    #     ]
    # }
    # print(all_data)
    # pass
    filename = "editable_format.xlsx"
    workbook = load_workbook("editable_format.xlsx")            # Loads the excel sheet
    worksheet = workbook.active                      # Select the active one 

    #  Updating the xlsx file here
    #  Making a new Curriculum Map
    #
    # Programme heading is stored --------------------------------------------------------------------
    row = 15
    column = 1

   
    worksheet.insert_rows(row)                          # Insert a row at 15
    cell_box = worksheet.cell(row=row,column=column)       # Create a cell
    cell_box.value = "Programmes"                    # Change the value of the cell to Programmes 
    cell_box.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    len_of_programmes  = len(all_data["programmes"])            # Length of data in the programmes array
    # We store the length 

    starting_character= 'B'            # Starting column  -- B
    char = ord(starting_character)          # Taking the order of the B
    location = starting_character + str(row)         # Adding 15 to B and storing in location
    for i in range(0, len_of_programmes):           # Looping through the array of programmes 

        char = ord(starting_character)
        # print("starting location:",location)
        TotalColumns += all_data["programmes"][i]["merge_columns"]
        val = char + all_data["programmes"][i]["merge_columns"] - 1            # Storing the merge column numbers + the order of that B Character and storing in val variable 

        new_location = val + 1
        new_location_character = chr(new_location)
        starting_character = chr(new_location)
        new_location = new_location_character + str(row)



        increment = chr(val)                                        # Increment varialbe to find character of that val variable 
        ending_column = increment + str(row)
        # print("ending column:",ending_column)

        merge_cells_columns = location + ':' + ending_column
        # print("Merge Cell Columns:",merge_cells_columns)

        worksheet.merge_cells(merge_cells_columns)       # Merging the columns with start row to row, start column to z and ending to ending_column 
        worksheet[location].value = all_data["programmes"][i]["name"]      # Getting each value of the array and storing in the location
        worksheet[location].alignment=Alignment(horizontal="center",wrap_text=True,shrink_to_fit=True,indent=0)
        set_border(worksheet, merge_cells_columns) 
        location = new_location                   # Updating location 
        # print("New Location:",location)


    # ---------------------------------------------------------------


    # Updating Modules name and code -----------------------------------


    row+=1
    column=1

    worksheet.insert_rows(row)                          # Insert a row at 15
    cell_box = worksheet.cell(row=row,column=column)       # Create a cell
    cell_box.value = "List programmes that this Curriculum Map refers to"                    # Change the value of the cell to Programmes 
    cell_box.alignment=Alignment(wrap_text=True,shrink_to_fit=True,indent=0)
    cell_box.border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    len_of_modules  = len(all_data["modules"])            # Length of data in the programmes array
    # We store the length 

    starting_character= 'B'            # Starting column  -- B
    char = ord(starting_character)          # Taking the order of the B
    location = starting_character + str(row)         # Adding 15 to B and storing in location
    for i in range(0, len_of_modules):           # Looping through the array of programmes 

        char = ord(starting_character)
        val = char + 1                              # Storing the merge column numbers + the order of that B Character and storing in val variable 
        new_location = val
        new_location_character = chr(new_location)
        starting_character = chr(new_location)
        new_location = new_location_character + str(row)



        increment = chr(val)                                        # Increment varialbe to find character of that val variable 

        # print("ending column:",ending_column)

        worksheet[location].value = all_data["modules"][i]["module_name"] + "\n" + all_data["modules"][i]["module_code"]      # Getting each value of the array and storing in the location
        worksheet[location].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
        worksheet[location].alignment=Alignment(horizontal="center",
                     vertical='top',
                     text_rotation=180,
                     wrap_text=True,
                     shrink_to_fit=True,
                     indent=0)

        location = new_location                   # Updating location 
        # print("New Location:",location)


    # --------------------------------------------------------------------

    # Updating table data ------------------------------------------------

    row+=1              # Next row 
    column=1            # Starting the column 
    len_table_data = len(all_data["table_data"])        # Length of table data i.e. 3

    
    total_columns = TotalColumns + 1           # Maximum columns 
    # print(total_columns)
    starting_character= 'A'                     # Starting from A
    char = ord(starting_character)                  # Taking the number of A
    val = char + total_columns -1                             # Storing the merge column numbers + the order of that B Character and storing in val variable 
    
    last_column = chr(val)


    location = starting_character + str(row)                # Adding 
    for i in range(0, len_table_data):           

        heading_merge_index = starting_character+str(row)+":"+last_column+str(row)




        worksheet.merge_cells(heading_merge_index)       # Merging the columns with start row to row, start column to z and ending to ending_column 
        worksheet[location].value = all_data["table_data"][i]["title"]      # Getting each value of the array and storing in the location
        color =all_data["table_data"][i]["bg_color"]
        color = color.replace('#', '')

        worksheet[location].fill = PatternFill(start_color=color, end_color=color,
                                        fill_type = "solid")
        set_border(worksheet, heading_merge_index)

        worksheet[location].alignment=Alignment(horizontal="center",wrap_text=True,shrink_to_fit=True,indent=0)

        len_of_data = len(all_data["table_data"][i]["data"])
        # print("Length of each record:",len_of_data)

        for j in range(0,len_of_data):
            row+=1
            starting_index = "A" 
            worksheet.insert_rows(row)             # Inserting new row 
            loc = starting_index +str(row)  # Location of the new row

            len_of_each_data = len(all_data["table_data"][i]["data"][j]["data"])
            print("Length of Each data:",len_of_each_data)

            worksheet[loc].value = all_data["table_data"][i]["data"][j]["heading"]
            worksheet[loc].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
            worksheet[loc].fill = PatternFill(start_color=color, end_color=color,
                                        fill_type = "solid")
            worksheet[loc].alignment=Alignment(horizontal="center",wrap_text=True,shrink_to_fit=True,indent=0)
            
            starting_index = "B" 
            loc = starting_index +str(row)  # Location of the new row

            for k in range(0,len_of_each_data):
            
                worksheet[loc].value = all_data["table_data"][i]["data"][j]["data"][k]
                worksheet[loc].border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
                worksheet[loc].fill = PatternFill(start_color=color, end_color=color,
                                        fill_type = "solid")
                char = ord(starting_index)                  # Taking the number of A
                val = char + 1                           # Storing the merge column numbers + the order of that B Character and storing in val variable 

                starting_index = chr(val)
                loc = starting_index +str(row)  # Location of the new row


        row+=1
        new_location = starting_character + str(row)

        location = new_location                   # Updating location 
        worksheet.insert_rows(row) 


        #--------------------------------------------------------------------



    workbook.save(savingFileName+".xlsx")

        # 15 Programs   (Heading- Programmes - and data )                   ---> DONE
    

        # 16 Modules    (Heading - List programmes that this Curriculum Map refers to - and data )  - DONE
        # 17 table Data 
            # - 17 title 
            # - 18A  title
            # - 18B onwards data 

        # Blue 
        # Green
        # Purple 
        # Peech
        # Light blue 


    # except:
    #     return Response(
    #         {
    #             "status": 400,
    #             "message": "Input field is missing"
    #         }
    #     )

    # ------------------------------


    file_to_read = open(savingFileName+'.xlsx', 'rb')

    response = HttpResponse(file_to_read,content_type='whatever')
    response['Content-Disposition'] = 'attachment; filename='+savingFileName+'.xlsx'



    # try:
    file_to_read = open(savingFileName+'.xlsx', 'rb')
    cFiles = CurriculumFiles.objects.create(filename = savingFileName+'.xlsx', excelFile = file_to_read.read())
    cFiles.save(savingFileName+'.xlsx')
    
    return HttpResponse(200)
    # except:
    #     return HttpResponseBadRequest(400)




